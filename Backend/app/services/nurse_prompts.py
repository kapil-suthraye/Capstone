from __future__ import annotations

from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from Backend.app.core.logging import logger
from Backend.app.models.nurse_prompt import NursePrompt


class NursePromptLoader:
    """
    Loads nurse prompts from the disease-specific sheets in the
    nurse_prompts_interqual.xlsx workbook.

    Sheet selection
    ---------------
    Only sheets whose names are NOT in EXCLUDED_SHEETS are read.
    The MASTER and INDEX sheets are excluded because they are
    administrative summaries, not source-of-truth prompt data.

    Each disease sheet (COVID-19, Sepsis, CAP, CHF-ADHF, …) has:
      - Row 0 : decorative title  (skipped)
      - Row 1 : column headers    (used to build the field map)
      - Subsequent rows : data rows, with occasional section-label
                          rows (no Prompt ID) that are skipped.

    The sheet name is used as the job_aid value since the disease
    sheets do not have a dedicated "Job Aid" column.
    """

    # Sheets that are NOT disease-diagnosis prompt sheets
    EXCLUDED_SHEETS: frozenset[str] = frozenset({
        "MASTER \u2014 All Prompts",   # unicode em-dash as stored in file
        "MASTER — All Prompts",        # fallback with regular dash
        "INDEX & LEGEND",
    })

    # Header aliases: maps the column name as it appears in the sheet
    # to the canonical field name used in NursePrompt
    HEADER_ALIASES: dict[str, str] = {
        # Standard names
        "Prompt ID":            "prompt_id",
        "Category":             "category",
        "Severity Level":       "severity_level",
        "Evaluation Prompt":    "evaluation_prompt",
        "Document Source":      "document_source",
        "Expected Finding":     "expected_finding",
        "Decision Impact":      "decision_impact",
        "RAG Search Keywords":  "rag_search_keywords",
        # Variant names found in disease sheets
        "Red Flag / Discrepancy":   "red_flag",
        "Red Flag":                 "red_flag",
        "InterQual / Guideline Ref": "guideline",
        "Guideline":                "guideline",
        "Job Aid":                  "job_aid",   # present in MASTER, not in disease sheets
    }

    def __init__(self, excel_path: str) -> None:
        self.excel_path = Path(excel_path)
        self.prompts: List[NursePrompt] = []
        self.prompt_lookup: Dict[str, NursePrompt] = {}
        self._load()

    # ------------------------------------------------------------------
    # Loading
    # ------------------------------------------------------------------

    def _load(self) -> None:
        if not self.excel_path.exists():
            raise FileNotFoundError(
                f"Nurse prompts workbook not found: {self.excel_path}"
            )

        workbook: Workbook = load_workbook(self.excel_path, data_only=True)

        disease_sheets = [
            name for name in workbook.sheetnames
            if name not in self.EXCLUDED_SHEETS
        ]

        if not disease_sheets:
            raise RuntimeError(
                f"No disease-diagnosis sheets found in '{self.excel_path.name}'. "
                f"All sheets present: {workbook.sheetnames}. "
                f"Check that EXCLUDED_SHEETS is not over-broad."
            )

        logger.info(
            f"Loading prompts from {len(disease_sheets)} disease sheet(s): "
            f"{disease_sheets}"
        )

        for sheet_name in disease_sheets:
            self._load_sheet(workbook[sheet_name], sheet_name)

        if not self.prompts:
            raise RuntimeError(
                f"Zero prompts loaded from disease sheets {disease_sheets}. "
                "Check that the sheets contain a 'Prompt ID' header row."
            )

        logger.info(f"Loaded {len(self.prompts)} prompts total.")

    def _load_sheet(self, sheet, sheet_name: str) -> None:
        """Parse one disease sheet and append its prompts to self.prompts."""
        rows = list(sheet.iter_rows(values_only=True))

        # Locate the header row — the row that contains "Prompt ID"
        header_row_idx: int | None = None
        for idx, row in enumerate(rows):
            if row and "Prompt ID" in row:
                header_row_idx = idx
                break

        if header_row_idx is None:
            logger.warning(
                f"Sheet '{sheet_name}': no 'Prompt ID' header found — skipping."
            )
            return

        # Build canonical field map from column index → field name
        raw_headers = rows[header_row_idx]
        field_map: dict[int, str] = {}
        for col_idx, cell_value in enumerate(raw_headers):
            if cell_value is None:
                continue
            canonical = self.HEADER_ALIASES.get(str(cell_value).strip())
            if canonical:
                field_map[col_idx] = canonical

        loaded = 0
        for row in rows[header_row_idx + 1:]:
            # Skip completely empty rows
            if not any(row):
                continue

            # Map each cell to its canonical field name
            fields: dict[str, str] = {}
            for col_idx, value in enumerate(row):
                field_name = field_map.get(col_idx)
                if field_name:
                    fields[field_name] = str(value).strip() if value is not None else ""

            prompt_id = fields.get("prompt_id", "")

            # Skip section-label rows (no Prompt ID, or Prompt ID looks like a heading)
            if not prompt_id or not self._looks_like_prompt_id(prompt_id):
                continue

            # Skip duplicate prompt IDs (sheet may overlap with another)
            if prompt_id in self.prompt_lookup:
                logger.debug(
                    f"Sheet '{sheet_name}': duplicate prompt_id '{prompt_id}' — skipping."
                )
                continue

            prompt = NursePrompt(
                prompt_id=prompt_id,
                # job_aid comes from the sheet name for disease sheets
                job_aid=fields.get("job_aid", sheet_name),
                category=fields.get("category", ""),
                severity_level=fields.get("severity_level") or None,
                evaluation_prompt=fields.get("evaluation_prompt", ""),
                document_source=fields.get("document_source") or None,
                expected_finding=fields.get("expected_finding") or None,
                red_flag=fields.get("red_flag") or None,
                guideline=fields.get("guideline") or None,
                decision_impact=fields.get("decision_impact") or None,
                rag_search_keywords=fields.get("rag_search_keywords") or None,
                sheet_name=sheet_name,
            )

            self.prompts.append(prompt)
            self.prompt_lookup[prompt_id] = prompt
            loaded += 1

        logger.info(f"Sheet '{sheet_name}': loaded {loaded} prompts.")

    @staticmethod
    def _looks_like_prompt_id(value: str) -> bool:
        """
        Returns False for section-label rows whose 'Prompt ID' cell contains
        a heading such as '▶  ACUTE LEVEL  — COVID-19' instead of an actual ID.
        A valid prompt ID is short (≤ 20 chars) and contains a hyphen.
        """
        return len(value) <= 20 and "-" in value

    # ------------------------------------------------------------------
    # Public API  (unchanged interface)
    # ------------------------------------------------------------------

    def get_all(self) -> List[NursePrompt]:
        return self.prompts

    def get_prompt(self, prompt_id: str) -> NursePrompt | None:
        return self.prompt_lookup.get(prompt_id)

    def get_job_aids(self) -> List[str]:
        return sorted({p.job_aid for p in self.prompts})

    def get_categories(self, job_aid: str | None = None) -> List[str]:
        prompts = self.prompts
        if job_aid:
            prompts = [p for p in prompts if p.job_aid == job_aid]
        return sorted({p.category for p in prompts})

    def get_by_job_aid(self, job_aid: str) -> List[NursePrompt]:
        return [p for p in self.prompts if p.job_aid == job_aid]

    def search(self, keyword: str) -> List[NursePrompt]:
        keyword = keyword.lower()
        return [p for p in self.prompts if keyword in p.evaluation_prompt.lower()]
